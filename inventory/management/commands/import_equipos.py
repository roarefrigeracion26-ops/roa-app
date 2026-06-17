import openpyxl
from django.core.management.base import BaseCommand
from inventory.models import Cliente, EquipoAA

EXCEL_PATH = r'C:\Users\Usuario\Downloads\Inventario de Equipos Olímpica S.A. (Respuestas).xlsx'

TIPO_EQUIPO_MAP = {
    'UMA': 'UMA',
    'Paquete': 'PAQUETE',
}

def parse_capacidad(val):
    if val is None:
        return ''
    return str(int(val)) if isinstance(val, float) and val == int(val) else str(val)

def parse_activo_fijo(val):
    if val is None or val == '' or val == 1:
        return ''
    return str(int(val)) if isinstance(val, float) and val == int(val) else str(val)


class Command(BaseCommand):
    help = 'Importa equipos desde el Excel de Olímpica'

    def handle(self, *args, **options):
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active

        creados = 0
        omitidos = 0
        errores = []

        for row in range(2, ws.max_row + 1):
            codigo = ws.cell(row, 1).value
            nombre_tienda = ws.cell(row, 2).value
            nombre = ws.cell(row, 3).value
            tipo_equipo_raw = ws.cell(row, 4).value
            ubicacion = ws.cell(row, 5).value or ''
            marca = ws.cell(row, 6).value or ''
            capacidad = parse_capacidad(ws.cell(row, 7).value)
            refrigerante = ws.cell(row, 8).value or ''
            tipo_correa = ws.cell(row, 9).value or ''
            modelo = ws.cell(row, 10).value or ''
            activo_fijo = parse_activo_fijo(ws.cell(row, 11).value)
            num_circuitos = ws.cell(row, 12).value or 1

            if not codigo or not nombre:
                continue

            tipo_equipo = TIPO_EQUIPO_MAP.get(tipo_equipo_raw, 'OTRO')

            try:
                cliente = Cliente.objects.get(codigo=codigo)
            except Cliente.DoesNotExist:
                errores.append(f'Fila {row}: Cliente con código {codigo} ({nombre_tienda}) no encontrado')
                continue

            if EquipoAA.objects.filter(cliente=cliente, nombre=nombre).exists():
                omitidos += 1
                continue

            id_qr = f'{nombre} {codigo}'

            EquipoAA.objects.create(
                cliente=cliente,
                nombre=nombre,
                id_qr=id_qr,
                tipo_equipo=tipo_equipo,
                ubicacion=ubicacion,
                marca=marca,
                modelo=modelo,
                capacidad=capacidad,
                refrigerante=refrigerante,
                tipo_correa=tipo_correa,
                activo_fijo=activo_fijo,
                num_circuitos=num_circuitos,
                activo=True,
            )
            creados += 1

        self.stdout.write(self.style.SUCCESS(f'Equipos creados: {creados}'))
        self.stdout.write(f'Equipos omitidos (ya existen): {omitidos}')
        if errores:
            self.stdout.write(self.style.ERROR('Errores:'))
            for e in errores:
                self.stdout.write(self.style.ERROR(f'  {e}'))
