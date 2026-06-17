import openpyxl
from django.core.management.base import BaseCommand
from inventory.models import Cliente, EquipoAA

EXCEL_PATH = r'C:\Users\Usuario\Downloads\Info. de Negocios & Equipos (Respuestas).xlsx'

TIPO_EQUIPO_MAP = {
    'split': 'SPLIT',
    'paquete': 'PAQUETE',
}

def normalizar_nombre(val):
    if val is None or val.strip() == '':
        return ''
    nombre = val.strip()
    if nombre.upper() == 'PISO-TECHO':
        return 'Piso Techo'
    return nombre.title()

class Command(BaseCommand):
    help = 'Importa equipos desde Info. de Negocios & Equipos'

    def handle(self, *args, **options):
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active

        creados = 0
        omitidos = 0
        errores = []

        for row in range(2, ws.max_row + 1):
            codigo = ws.cell(row, 1).value
            nombre_tienda = ws.cell(row, 2).value
            tipo_equipo_raw = ws.cell(row, 3).value
            nombre_base_raw = ws.cell(row, 4).value
            marca = ws.cell(row, 5).value or ''
            cantidad_raw = ws.cell(row, 6).value

            if not codigo or not tipo_equipo_raw:
                continue

            try:
                codigo = int(float(str(codigo)))
            except (ValueError, TypeError):
                errores.append(f'Fila {row}: Codigo invalido "{codigo}"')
                continue

            try:
                cantidad = int(float(str(cantidad_raw)))
            except (ValueError, TypeError):
                cantidad = 1

            tipo_equipo = TIPO_EQUIPO_MAP.get(tipo_equipo_raw.strip().lower(), 'OTRO')
            base_name = normalizar_nombre(nombre_base_raw)

            if not base_name:
                base_name = tipo_equipo_raw.strip().title()

            try:
                cliente = Cliente.objects.get(codigo=codigo)
            except Cliente.DoesNotExist:
                errores.append(f'Fila {row}: Cliente con codigo {codigo} ({nombre_tienda}) no encontrado')
                continue

            for i in range(1, cantidad + 1):
                nombre = f'{base_name} {i}'
                id_qr = f'{nombre} {codigo}'

                if EquipoAA.objects.filter(cliente=cliente, nombre=nombre).exists():
                    omitidos += 1
                    continue

                EquipoAA.objects.create(
                    cliente=cliente,
                    nombre=nombre,
                    id_qr=id_qr,
                    tipo_equipo=tipo_equipo,
                    ubicacion='',
                    marca=marca,
                    modelo='',
                    capacidad='',
                    refrigerante='',
                    tipo_correa='',
                    activo_fijo='',
                    num_circuitos=1,
                    voltaje='',
                    fases='',
                    activo=True,
                )
                creados += 1

        self.stdout.write(self.style.SUCCESS(f'Equipos creados: {creados}'))
        self.stdout.write(f'Equipos omitidos (ya existen): {omitidos}')
        if errores:
            self.stdout.write(self.style.ERROR('Errores:'))
            for e in errores:
                self.stdout.write(self.style.ERROR(f'  {e}'))
