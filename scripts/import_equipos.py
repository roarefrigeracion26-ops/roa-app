"""
Standalone import script for EquipoAA from Excel to Supabase.
Usage: python scripts/import_equipos.py "postgres://user:pass@host:5432/dbname?sslmode=require"
"""

import sys
import openpyxl
from sqlalchemy import create_engine, text

EXCEL_PATH = r'C:\Users\Usuario\Downloads\Inventario de Equipos Olímpica S.A. (Respuestas).xlsx'

TIPO_EQUIPO_MAP = {
    'UMA': 'UMA',
    'Paquete': 'PAQUETE',
}


def parse_capacidad(val):
    if val is None:
        return ''
    return str(int(val)) if isinstance(val, float) and val == int(val) else str(val)


def parse_activo_fijo(val):
    if val is None or val == '' or val == 1:
        return ''
    return str(int(val)) if isinstance(val, float) and val == int(val) else str(val)


def main():
    if len(sys.argv) < 2:
        print('Usage: python scripts/import_equipos.py "DATABASE_URL"')
        sys.exit(1)

    db_url = sys.argv[1]
    engine = create_engine(db_url)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    creados = 0
    omitidos = 0
    errores = []

    with engine.begin() as conn:
        for row in range(2, ws.max_row + 1):
            codigo = ws.cell(row, 1).value
            nombre_tienda = ws.cell(row, 2).value
            nombre = ws.cell(row, 3).value
            tipo_equipo_raw = ws.cell(row, 4).value
            ubicacion = ws.cell(row, 5).value or ''
            marca = ws.cell(row, 6).value or ''
            capacidad = parse_capacidad(ws.cell(row, 7).value)
            refrigerante = ws.cell(row, 8).value or ''
            tipo_correa = ws.cell(row, 9).value or ''
            modelo = ws.cell(row, 10).value or ''
            activo_fijo = parse_activo_fijo(ws.cell(row, 11).value)
            num_circuitos = ws.cell(row, 12).value or 1

            if not codigo or not nombre:
                continue

            tipo_equipo = TIPO_EQUIPO_MAP.get(tipo_equipo_raw, 'OTRO')

            result = conn.execute(
                text('SELECT id FROM inventory_cliente WHERE codigo = :codigo'),
                {'codigo': codigo}
            )
            cliente_row = result.fetchone()
            if not cliente_row:
                errores.append(f'Fila {row}: Cliente con código {codigo} ({nombre_tienda}) no encontrado')
                continue

            cliente_id = cliente_row[0]

            exists = conn.execute(
                text('SELECT 1 FROM inventory_equipoaa WHERE cliente_id = :cliente_id AND nombre = :nombre'),
                {'cliente_id': cliente_id, 'nombre': nombre}
            ).fetchone()
            if exists:
                omitidos += 1
                continue

            id_qr = f'{nombre} {codigo}'

            conn.execute(
                text("""
                    INSERT INTO inventory_equipoaa
                        (cliente_id, nombre, tipo_equipo, ubicacion, marca, modelo,
                         capacidad, refrigerante, tipo_correa, activo_fijo,
                         num_circuitos, activo, voltaje, fases, id_qr)
                    VALUES
                        (:cliente_id, :nombre, :tipo_equipo, :ubicacion, :marca, :modelo,
                         :capacidad, :refrigerante, :tipo_correa, :activo_fijo,
                         :num_circuitos, :activo, :voltaje, :fases, :id_qr)
                """),
                {
                    'cliente_id': cliente_id,
                    'nombre': nombre,
                    'id_qr': id_qr,
                    'tipo_equipo': tipo_equipo,
                    'ubicacion': ubicacion,
                    'marca': marca,
                    'modelo': modelo,
                    'capacidad': capacidad,
                    'refrigerante': refrigerante,
                    'tipo_correa': tipo_correa,
                    'activo_fijo': activo_fijo,
                    'num_circuitos': num_circuitos,
                    'activo': True,
                    'voltaje': '',
                    'fases': '',
                }
            )
            creados += 1

    print(f'Equipos creados: {creados}')
    print(f'Equipos omitidos (ya existen): {omitidos}')
    if errores:
        print('Errores:')
        for e in errores:
            print(f'  {e}')


if __name__ == '__main__':
    main()
